"""Extract Cummins price list PDF (custom-encoded fonts) to Excel via cell OCR."""

from __future__ import annotations

import argparse
import re
import sys
import time
from pathlib import Path

import fitz
import numpy as np
import openpyxl
import pdfplumber
from openpyxl.utils import get_column_letter

try:
    import easyocr
except ImportError:
    print("Install dependencies: pip install pymupdf pdfplumber openpyxl easyocr")
    sys.exit(1)

COLUMNS = [
    "Sr_No",
    "Part_No",
    "Description",
    "LP",
    "MRP",
    "HSN_Code",
    "GST",
]

SCALE = 1.25
CHECKPOINT_EVERY = 25

HEADER_MARKERS = (
    "part_number",
    "part description",
    "cil dbu parts price revision",
    "suggested list price",
    "mrp price",
)


def clean_digits(value: str, keep_dot: bool = False) -> str:
    pattern = r"[^0-9.]" if keep_dot else r"[^0-9]"
    return re.sub(pattern, "", value or "")


def clean_part_no(value: str) -> str:
    return clean_digits(value)


def clean_gst(value: str) -> str:
    match = re.search(r"\b(18|28|12|5)\b", value or "")
    if match:
        return f"{match.group(1)}%"
    match = re.search(r"(18|28|12|5)", value or "")
    return f"{match.group(1)}%" if match else (value or "").strip()


def normalize_row(values: list[str]) -> dict[str, str]:
    row = {col: (values[i] if i < len(values) else "").strip() for i, col in enumerate(COLUMNS)}

    row["Part_No"] = clean_part_no(row["Part_No"])
    row["Description"] = re.sub(r"\s+", " ", row["Description"]).strip()
    row["LP"] = clean_digits(row["LP"], keep_dot=True)
    row["MRP"] = clean_digits(row["MRP"])
    row["HSN_Code"] = clean_digits(row["HSN_Code"])
    row["Sr_No"] = clean_digits(row["Sr_No"])
    row["GST"] = clean_gst(row["GST"])

    # Move 8-digit HSN values that landed in MRP/LP columns.
    for source in ("MRP", "LP"):
        if len(row[source]) == 8 and row[source].isdigit() and not row["HSN_Code"]:
            row["HSN_Code"] = row[source]
            row[source] = ""

    if row["LP"] and "." not in row["LP"] and not row["MRP"] and row["HSN_Code"]:
        if len(row["LP"]) <= 5:
            row["MRP"] = row["LP"]
            row["LP"] = ""

    if not row["GST"]:
        row["GST"] = "18%"

    return row


def is_valid_row(row: dict[str, str]) -> bool:
    if not re.fullmatch(r"\d{3,8}", row.get("Part_No", "")):
        return False

    blob = " ".join(row.values()).lower()
    return not any(marker in blob for marker in HEADER_MARKERS)


def ocr_cell(reader, image: np.ndarray, bbox, scale: float, column: str) -> str:
    x0, top, x1, bottom = bbox
    sx0 = max(0, int(x0 * scale))
    sy0 = max(0, int(top * scale))
    sx1 = int(x1 * scale)
    sy1 = int(bottom * scale)

    if sy1 <= sy0 or sx1 <= sx0:
        return ""

    crop = image[sy0:sy1, sx0:sx1]
    if crop.size == 0:
        return ""

    kwargs: dict = {"detail": 0, "paragraph": False}
    if column == "GST":
        kwargs["allowlist"] = "0123456789%"
    elif column == "Part_No":
        kwargs["allowlist"] = "0123456789"
    elif column in {"Sr_No", "LP", "MRP", "HSN_Code"}:
        kwargs["allowlist"] = "0123456789."

    return " ".join(reader.readtext(crop, **kwargs)).strip()


def extract_page(
    page_index: int,
    plumber_page,
    reader,
    image: np.ndarray,
    scale: float,
    sr_counter: int,
) -> tuple[list[list[str]], int]:
    rows_out: list[list[str]] = []
    tables = plumber_page.find_tables()
    if not tables:
        return rows_out, sr_counter

    for table_row in tables[0].rows:
        raw_values: list[str] = []
        for column, cell in zip(COLUMNS, table_row.cells):
            if cell is None:
                raw_values.append("")
            else:
                raw_values.append(ocr_cell(reader, image, cell, scale, column))

        row = normalize_row(raw_values)
        if not is_valid_row(row):
            continue

        if not row["Sr_No"]:
            sr_counter += 1
            row["Sr_No"] = str(sr_counter)
        else:
            sr_counter = max(sr_counter, int(row["Sr_No"]))

        rows_out.append([row[col] for col in COLUMNS])

    return rows_out, sr_counter


def autosize_columns(ws) -> None:
    for col_idx, header in enumerate(COLUMNS, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def save_workbook(path: Path, rows: list[list[str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price List"
    ws.append(COLUMNS)
    for row in rows:
        ws.append(row)
    autosize_columns(ws)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract Cummins PDF price list to Excel")
    parser.add_argument(
        "--pdf",
        default=r"c:\Users\user\Downloads\Cummins Price List-LP-MRP.pdf",
        help="Path to source PDF",
    )
    parser.add_argument(
        "--output",
        default=r"c:\Users\user\Downloads\Cummins_Price_List.xlsx",
        help="Path to output Excel file",
    )
    parser.add_argument("--start-page", type=int, default=1, help="1-based start page")
    parser.add_argument("--end-page", type=int, default=0, help="1-based end page (0 = all)")
    args = parser.parse_args()

    pdf_path = Path(args.pdf)
    output_path = Path(args.output)
    checkpoint_path = output_path.with_suffix(".checkpoint.xlsx")

    if not pdf_path.exists():
        print(f"PDF not found: {pdf_path}", flush=True)
        sys.exit(1)

    print("Loading OCR model (first run downloads models)...", flush=True)
    reader = easyocr.Reader(["en"], gpu=False, verbose=False)

    fitz_doc = fitz.open(pdf_path)
    total_pages = fitz_doc.page_count
    start = max(1, args.start_page) - 1
    end = total_pages if args.end_page <= 0 else min(args.end_page, total_pages)

    all_rows: list[list[str]] = []
    sr_counter = 0
    started = time.time()

    print(f"Extracting pages {start + 1} to {end} from {pdf_path.name}", flush=True)
    print(f"Output: {output_path}", flush=True)
    print(f"Estimated time: ~{(end - start) * 85 / 3600:.1f} hours", flush=True)

    with pdfplumber.open(pdf_path) as pdf:
        for page_index in range(start, end):
            page_num = page_index + 1
            t0 = time.time()

            fitz_page = fitz_doc[page_index]
            pix = fitz_page.get_pixmap(matrix=fitz.Matrix(SCALE, SCALE))
            image = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
            if pix.n == 4:
                image = image[:, :, :3]

            page_rows, sr_counter = extract_page(
                page_index,
                pdf.pages[page_index],
                reader,
                image,
                SCALE,
                sr_counter,
            )
            all_rows.extend(page_rows)

            elapsed = time.time() - t0
            done = page_index - start + 1
            total = end - start
            avg = (time.time() - started) / done
            remaining = avg * (total - done)
            print(
                f"Page {page_num}/{end}: {len(page_rows)} rows | "
                f"{elapsed:.1f}s | total rows: {len(all_rows)} | ETA: {remaining/3600:.1f} h",
                flush=True,
            )

            if done % CHECKPOINT_EVERY == 0:
                save_workbook(checkpoint_path, all_rows)
                print(f"  Checkpoint saved: {checkpoint_path}", flush=True)

    fitz_doc.close()
    save_workbook(output_path, all_rows)
    print(f"\nDone. {len(all_rows)} rows saved to {output_path}", flush=True)
    if checkpoint_path.exists() and checkpoint_path != output_path:
        checkpoint_path.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
